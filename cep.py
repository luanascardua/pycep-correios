from pycep_correios.client import WebService, get_address_from_cep
from pycep_correios import exceptions as correiosExceptions
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl import load_workbook
from openpyxl.styles import Font
from tkinter import filedialog
import os

class Cep:

    def __init__(self, file='cep.xlsx'):
        self.file = file

        if not os.path.isfile(self.file):
            self.file = filedialog.askopenfilename(title='Selecione a planilha')
            extension = os.path.splitext(self.file)
        try:
            self.wb = load_workbook(self.file)
        except InvalidFileException:
            print(f'\nArquivo {extension[1]} inválido. Selecione um arquivo nos formatos suportados: '
                  f'.xlsx,.xlsm,.xltx,.xltm')
            os._exit(1)
        
        self.ws = self.wb.active
        
        if self.ws.cell(row=1, column=6).value == None:
            self.ws.cell(row=1, column=6).value = 'status'
            self.wb.save(self.file)

        column_status = {
            line.value for line in self.ws['F']
            if line.value != None
        }
        self.first_line = len(column_status) + 1
        self.last_line = len(self.ws['A']) + 1
        

    def read_data(self, line):
        self.cep = str(self.ws.cell(row=line, column=1).value)
        self.cep = self.cep[:5] + '-' + self.cep[5:]

    def insert_data(self, line):
        style_column_name = Font(name='Calibri',
                    size = 11,
                    bold = True)
        b1 = self.ws['B1']
        c1 = self.ws['C1']
        d1 = self.ws['D1']
        e1 = self.ws['E1']
        f1 = self.ws['F1']
        b1.font = style_column_name
        c1.font = style_column_name
        d1.font = style_column_name
        e1.font = style_column_name
        f1.font = style_column_name
        
        if self.ws.cell(row=1, column=2).value == None:
            self.ws.cell(row=1, column=2).value = 'bairro'
            self.ws.cell(row=1, column=3).value = 'cidade'
            self.ws.cell(row=1, column=4).value = 'logradouro'
            self.ws.cell(row=1, column=5).value = 'estado'

        if self.status == 'CEP encontrado':
            self.ws.cell(row=line, column=2).value = self.endereco['bairro']
            self.ws.cell(row=line, column=3).value = self.endereco['cidade']
            self.ws.cell(row=line, column=4).value = self.endereco['logradouro']
            self.ws.cell(row=line, column=5).value = self.endereco['uf']
        self.ws.cell(row=line, column=6).value = self.status

        self.wb.save(self.file)

    def search_cep(self):
        try:
            self.endereco = get_address_from_cep((self.cep), webservice=WebService.VIACEP)
            self.status = 'CEP encontrado'
        except correiosExceptions.InvalidCEP:
            self.status = 'CEP Inválido'
        except correiosExceptions.CEPNotFound:
            self.status = 'CEP inexistente'
